#!/usr/bin/env python3
"""
CPM-TO-P6 Bridge  —  COSEDIL S.p.A. · PMO
Replica la logica del file Excel "CPM to P6" in Python.

Comandi:
  analizza   — mostra RIEPILOGO e ALERT senza scrivere
  rigenera   — riscrive BRIDGE_SIL e P6_IMPORT_PULITO nell'Excel
  export     — esporta P6_IMPORT_PULITO come CSV pronto per import in P6

Uso:
  python cpm_to_p6.py analizza
  python cpm_to_p6.py rigenera
  python cpm_to_p6.py export --output p6_import.csv
  python cpm_to_p6.py analizza --sil 14    # forza SIL corrente
  python cpm_to_p6.py analizza --file "altro_bridge.xlsx"
"""

import sys
import argparse
import csv
import datetime
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import io
# Fix encoding console Windows (cp1252 non supporta caratteri Unicode)
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf-8-sig"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Installa openpyxl:  pip install openpyxl")

# ─── Percorso default del file Bridge ────────────────────────────────────────
DEFAULT_FILE = Path(__file__).parent / "file" / "CPM to P6_rev.0.13live.xlsx"

# ─── Costanti colonne (0-based) ───────────────────────────────────────────────
# SIL diretti
SD_WBS   = 0   # Cod. WBS
SD_SIL   = 3   # Cod. S.I.L.
SD_DATA  = 4   # Data
SD_ART   = 6   # Articolo
SD_QTA   = 10  # Quantità
SD_IMP   = 11  # Importo (ricavo)

# SIL indiretti
SI_SIL   = 1   # Cod. S.I.L.
SI_DATA  = 2   # Data
SI_ART   = 4   # Articolo
SI_QTA   = 8   # Quantità
SI_IMP   = 9   # Importo
SI_WBS   = 13  # Cod. WBS

# MAPPING
MP_WBS   = 7   # Cod. WBS
MP_DES   = 8   # Des. WBS
MP_TIPO  = 9   # Cod. TIPONODO  (DI / IN)
MP_ACTS  = 11  # Activity ID P6 (comma-separated or "—" / "?")

# BUDGET da CPM  (header a riga index 3)
BD_ART   = 4   # Articolo
BD_COST  = 12  # Costo unitario

# INPUT da P6 — XER TASK  (indici nel %F header, 0-based includendo "%F" in pos 0)
XER_ID   = 14  # task_code
XER_NM   = 15  # task_name
XER_PCT  = 5   # phys_complete_pct
XER_DUR  = 23  # target_drtn_hr_cnt


# ─── Caricamento dati ─────────────────────────────────────────────────────────

def _str(v) -> str:
    return "" if v is None else str(v).strip()


def load_mapping(ws) -> Tuple[Dict, Dict, Dict]:
    """
    Restituisce:
      wbs_to_acts  : {wbs: [act_id, ...]}  (stringhe)
      wbs_to_des   : {wbs: descrizione}
      wbs_to_tipo  : {wbs: "DI" | "IN"}
    """
    wbs_to_acts: Dict[str, List[str]] = {}
    wbs_to_des: Dict[str, str] = {}
    wbs_to_tipo: Dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        wbs  = _str(row[MP_WBS])
        des  = _str(row[MP_DES])
        tipo = _str(row[MP_TIPO])
        acts = _str(row[MP_ACTS])
        if not wbs:
            continue
        if wbs not in wbs_to_des:
            wbs_to_des[wbs] = des
        if wbs not in wbs_to_tipo:
            wbs_to_tipo[wbs] = tipo
        if acts and acts not in ("—", "—", "?"):  # em-dash or normal dash
            for a in acts.split(","):
                a = a.strip()
                if a:
                    wbs_to_acts.setdefault(wbs, [])
                    if a not in wbs_to_acts[wbs]:
                        wbs_to_acts[wbs].append(a)

    return wbs_to_acts, wbs_to_des, wbs_to_tipo


def load_budget(ws) -> Dict[str, float]:
    """Restituisce {articolo: costo_unitario}"""
    art_to_cost: Dict[str, float] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):  # header a riga 4
        art  = _str(row[BD_ART])
        cost = row[BD_COST]
        if art and art not in art_to_cost:
            try:
                art_to_cost[art] = float(cost) if cost is not None else 0.0
            except (TypeError, ValueError):
                art_to_cost[art] = 0.0
    return art_to_cost


def load_p6_tasks(ws) -> Dict[str, Dict]:
    """
    Legge la sezione TASK del file XER (foglio INPUT da P6).
    Restituisce {task_code_str: {name, pct, dur}}
    """
    act_info: Dict[str, Dict] = {}
    rows = list(ws.iter_rows(values_only=True))

    # Trova la sezione %T TASK (l'ultima, quella vera con tutte le attività)
    task_t_row = -1
    for i, row in enumerate(rows):
        if _str(row[0]) == "%T" and _str(row[1]) == "TASK":
            task_t_row = i

    if task_t_row < 0:
        return act_info

    # Riga %F = header dei campi
    hdr_row = task_t_row + 1
    if hdr_row >= len(rows) or _str(rows[hdr_row][0]) != "%F":
        return act_info

    headers = [_str(v) for v in rows[hdr_row]]
    try:
        ci_id  = headers.index("task_code")
        ci_nm  = headers.index("task_name")
        ci_pct = headers.index("phys_complete_pct")
        ci_dur = headers.index("target_drtn_hr_cnt")
    except ValueError as e:
        print(f"[WARN] Campo XER mancante: {e}")
        ci_id  = XER_ID
        ci_nm  = XER_NM
        ci_pct = XER_PCT
        ci_dur = XER_DUR

    for row in rows[hdr_row + 1:]:
        if _str(row[0]) != "%R":
            if _str(row[0]) == "%T":
                break  # prossima sezione
            continue
        act_id = _str(row[ci_id]) if ci_id < len(row) else ""
        name   = _str(row[ci_nm]) if ci_nm < len(row) else ""
        pct    = row[ci_pct] if ci_pct < len(row) else 0
        dur    = row[ci_dur] if ci_dur < len(row) else 0
        if not act_id:
            continue
        try:
            pct_f = float(pct) if pct is not None else 0.0
        except (TypeError, ValueError):
            pct_f = 0.0
        try:
            dur_f = float(dur) if dur is not None else 0.0
        except (TypeError, ValueError):
            dur_f = 0.0
        act_info[act_id] = {"name": name, "pct": pct_f, "dur": dur_f}

    return act_info


def load_sil_diretti(ws) -> List[Dict]:
    """Restituisce lista di record SIL diretti."""
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        wbs = _str(row[SD_WBS])
        sil = row[SD_SIL]
        if not wbs or sil is None:
            continue
        records.append({
            "wbs":  wbs,
            "sil":  int(sil) if sil is not None else 0,
            "data": row[SD_DATA],
            "art":  _str(row[SD_ART]),
            "qta":  float(row[SD_QTA]) if row[SD_QTA] is not None else 0.0,
            "imp":  float(row[SD_IMP]) if row[SD_IMP] is not None else 0.0,
            "tipo_base": "Diretto",
        })
    return records


def load_sil_indiretti(ws) -> List[Dict]:
    """Restituisce lista di record SIL indiretti."""
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        wbs = _str(row[SI_WBS])
        sil = row[SI_SIL]
        if not wbs or sil is None:
            continue
        qta = float(row[SI_QTA]) if row[SI_QTA] is not None else 0.0
        imp = float(row[SI_IMP]) if row[SI_IMP] is not None else 0.0
        if qta == 0 and imp == 0:
            continue
        records.append({
            "wbs":  wbs,
            "sil":  int(sil) if sil is not None else 0,
            "data": row[SI_DATA],
            "art":  _str(row[SI_ART]),
            "qta":  qta,
            "imp":  imp,
            "tipo_base": "Indiretto",
        })
    return records


# ─── Logica core: esplosione SIL → BRIDGE_SIL ────────────────────────────────

def _compute_weights(acts: List[str], act_info: Dict) -> List[float]:
    """Calcola i pesi proporzionali per distribuire il costo alle attività."""
    pcts = [act_info.get(a, {}).get("pct", 0.0) for a in acts]
    sum_pct = sum(pcts)
    if sum_pct > 0:
        return [p / sum_pct for p in pcts]

    durs = [act_info.get(a, {}).get("dur", 0.0) for a in acts]
    sum_dur = sum(durs)
    if sum_dur > 0:
        return [d / sum_dur for d in durs]

    return [1.0 / len(acts)] * len(acts)


def generate_bridge(
    sil_records: List[Dict],
    wbs_to_acts: Dict,
    wbs_to_des: Dict,
    wbs_to_tipo: Dict,
    act_info: Dict,
    art_to_cost: Dict,
) -> List[List]:
    """
    Genera le righe del BRIDGE_SIL.
    Schema colonne (0-based):
      0  SIL #
      1  Data SIL
      2  Cod. WBS CPM
      3  Des. WBS CPM
      4  Articolo CPM
      5  Importo SIL Ricavo (€)
      6  P6 Activity ID
      7  P6 Activity Name
      8  Peso Proporzionale
      9  Importo Distribuito Ricavo (€)
      10 Importo Cumulativo Ricavo (€)   ← calcolato dopo
      11 Tipo
      12 Costo Distribuito (€)
    """
    out: List[List] = []

    for rec in sil_records:
        wbs      = rec["wbs"]
        sil_n    = rec["sil"]
        data     = rec["data"]
        art      = rec["art"]
        qta      = rec["qta"]
        imp      = rec["imp"]  # ricavo
        tipo_base = rec["tipo_base"]

        costo_unit = art_to_cost.get(art, 0.0)
        costo_tot  = qta * costo_unit
        des_wbs    = wbs_to_des.get(wbs, "")
        tipo_map   = wbs_to_tipo.get(wbs, "")
        acts       = wbs_to_acts.get(wbs)

        if not acts:
            # Nessun mapping P6
            if tipo_map == "IN":
                tipo_final = "Indiretto"
            elif wbs == "SIC":
                tipo_final = "Sicurezza"
            else:
                tipo_final = "MAPPING MANCANTE"
            out.append([
                sil_n, data, wbs, des_wbs, art,
                imp, "N/A", "MAPPING MANCANTE",
                1.0, imp, 0.0, tipo_final, costo_tot,
            ])
            continue

        pesi = _compute_weights(acts, act_info)
        for act_id, w in zip(acts, pesi):
            act_name = act_info.get(act_id, {}).get("name", "")
            out.append([
                sil_n, data, wbs, des_wbs, art,
                imp, act_id, act_name,
                round(w, 6),
                round(imp * w, 6),
                0.0,  # cumulativo: calcolato sotto
                tipo_base,
                round(costo_tot * w, 6),
            ])

    # ── Calcolo cumulativo per Activity ID (ordinato per act_id, data) ──
    out.sort(key=lambda r: (str(r[6]), r[1] if isinstance(r[1], (int, float)) else 0))
    cur_act = None
    cum = 0.0
    for row in out:
        act = str(row[6])
        if act != cur_act:
            cur_act = act
            cum = 0.0
        cum += row[9]
        row[10] = round(cum, 6)

    return out


# ─── Generazione P6_IMPORT_PULITO ─────────────────────────────────────────────

def generate_p6_import(
    bridge_rows: List[List],
    sil_corrente: int,
    act_info: Dict,
) -> List[List]:
    """
    Genera le righe dati di P6_IMPORT_PULITO.
    Colonne: Activity ID | Activity Name | Actual This Period Cost | Actual Total Cost
    """
    # Aggrega per Activity ID
    total: Dict[str, float] = defaultdict(float)
    periodo: Dict[str, float] = defaultdict(float)
    names: Dict[str, str] = {}

    for row in bridge_rows:
        act_id   = str(row[6])
        imp_dist = float(row[9])   # ricavo distribuito
        sil_n    = int(row[0])

        total[act_id] += imp_dist
        if sil_n == sil_corrente:
            periodo[act_id] += imp_dist
        if act_id not in names:
            names[act_id] = row[7] or act_info.get(act_id, {}).get("name", "")

    # Tutti gli act_id presenti nell'XER (anche quelli senza costo)
    all_acts = sorted(
        set(list(act_info.keys()) + list(total.keys())),
        key=lambda x: (len(x), x)
    )

    result = []
    for act_id in all_acts:
        act_name   = names.get(act_id) or act_info.get(act_id, {}).get("name", "")
        this_per   = round(periodo.get(act_id, 0.0), 6)
        act_total  = round(total.get(act_id, 0.0), 6)
        result.append([act_id, act_name, this_per, act_total])

    return result


# ─── Report RIEPILOGO ─────────────────────────────────────────────────────────

def show_riepilogo(sil_records: List[Dict], bridge_rows: List[List]) -> None:
    """Mostra la quadratura WBS: Totale SIL vs Totale Bridge P6."""
    sil_by_wbs: Dict[str, float] = defaultdict(float)
    des_by_wbs: Dict[str, str]   = {}
    cnt_by_wbs: Dict[str, int]   = defaultdict(int)

    for rec in sil_records:
        wbs = rec["wbs"]
        sil_by_wbs[wbs] += rec["imp"]
        cnt_by_wbs[wbs] += 1
        if wbs not in des_by_wbs:
            des_by_wbs[wbs] = ""

    bridge_by_wbs: Dict[str, float] = defaultdict(float)
    for row in bridge_rows:
        wbs = str(row[2])
        bridge_by_wbs[wbs] += float(row[9])

    all_wbs = sorted(set(list(sil_by_wbs.keys()) + list(bridge_by_wbs.keys())))

    print()
    print("═" * 100)
    print(" RIEPILOGO — Quadratura WBS: SIL CPM vs Bridge P6")
    print("═" * 100)
    fmt = "{:<8}  {:>4}  {:>16}  {:>16}  {:>12}  {:>8}  {}"
    print(fmt.format("WBS", "N°", "Tot SIL (€)", "Tot Bridge (€)", "Delta (€)", "Delta %", "Check"))
    print("─" * 100)

    tot_sil = tot_bridge = 0.0
    tot_cnt = 0

    for wbs in all_wbs:
        s = sil_by_wbs.get(wbs, 0.0)
        b = bridge_by_wbs.get(wbs, 0.0)
        d = s - b
        pct = (d / s * 100) if s else 0.0
        check = "✓ OK" if abs(d) <= 1.0 else "⚠ DELTA"
        n = cnt_by_wbs.get(wbs, 0)
        print(fmt.format(
            wbs, n,
            f"{s:,.2f}", f"{b:,.2f}", f"{d:,.2f}",
            f"{pct:.3f}%", check
        ))
        tot_sil    += s
        tot_bridge += b
        tot_cnt    += n

    print("─" * 100)
    d_tot = tot_sil - tot_bridge
    pct_tot = (d_tot / tot_sil * 100) if tot_sil else 0.0
    chk_tot = "✓ QUADRA" if abs(d_tot) <= 1.0 else "⚠ NON QUADRA"
    print(fmt.format(
        "TOTALE", tot_cnt,
        f"{tot_sil:,.2f}", f"{tot_bridge:,.2f}", f"{d_tot:,.2f}",
        f"{pct_tot:.3f}%", chk_tot
    ))
    print()


# ─── Report ALERT ─────────────────────────────────────────────────────────────

def show_alert(
    bridge_rows: List[List],
    act_info: Dict,
    sil_corrente: int,
) -> None:
    """Elenca le attività con costo > 0 nel SIL corrente ma % fisica = 0."""
    # Aggrega per attività nel SIL corrente
    periodo_costo: Dict[str, float] = defaultdict(float)
    names: Dict[str, str] = {}

    for row in bridge_rows:
        if int(row[0]) != sil_corrente:
            continue
        act_id = str(row[6])
        periodo_costo[act_id] += float(row[9])
        if act_id not in names:
            names[act_id] = row[7] or ""

    alerts = []
    for act_id, costo in periodo_costo.items():
        if costo <= 0:
            continue
        info = act_info.get(act_id, {})
        pct  = info.get("pct", 0.0)
        dur  = info.get("dur", 0.0)
        if pct == 0:
            metodo = "DUR_FALLBACK" if dur > 0 else "NESSUN_PESO"
            alerts.append((act_id, names[act_id], costo, pct, dur, metodo))

    print()
    print("═" * 95)
    print(f" ALERT — SIL {sil_corrente} — Attività con ricavo > 0 ma % fisica = 0")
    print("═" * 95)
    if not alerts:
        print("  Nessun alert. Tutte le attività con costo hanno % fisica > 0.")
    else:
        fmt = "{:<8}  {:>12}  {:>8}  {:>8}  {:16}  {}"
        print(fmt.format("Act ID", "Ricavo (€)", "% Fisica", "Durata", "Metodo", "Nome attività"))
        print("─" * 95)
        for act_id, name, costo, pct, dur, metodo in sorted(alerts):
            print(fmt.format(act_id, f"{costo:,.2f}", f"{pct:.0f}%", f"{dur:.0f}h", metodo, name[:45]))
    print()


# ─── Scrittura Excel ──────────────────────────────────────────────────────────

BRIDGE_HEADERS = [
    "SIL #", "Data SIL", "Cod. WBS CPM", "Des. WBS CPM", "Articolo CPM",
    "Importo SIL Ricavo (€)", "P6 Activity ID", "P6 Activity Name",
    "Peso Proporzionale", "Importo Distribuito Ricavo (€)",
    "Importo Cumulativo Ricavo (€)", "Tipo", "Costo Distribuito (€)",
]

P6_IMPORT_HEADERS = [
    "Activity ID", "Activity Name",
    "Actual This Period Cost", "Actual Total Cost",
]


def _write_bridge(ws, bridge_rows: List[List]) -> None:
    """Scrive BRIDGE_SIL nell'Excel (sovrascrive dati, conserva intestazione)."""
    # Cancella dati esistenti (riga 2 in poi)
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

    # Scrivi headers
    for c, h in enumerate(BRIDGE_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)

    # Scrivi dati
    for r, row in enumerate(bridge_rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)


def _write_p6_import(ws, import_rows: List[List], sil_corrente: int, sil_max: int) -> None:
    """Aggiorna il foglio P6_IMPORT_PULITO."""
    # Aggiorna SIL corrente in B4 e C4
    ws["B4"] = sil_corrente
    ws["C4"] = sil_max

    # Calcola totali periodo / cumulativo per il summary in B6
    tot_periodo   = sum(r[2] for r in import_rows)
    tot_cumulativo = sum(r[3] for r in import_rows)
    ws["A6"] = (
        f"SIL {sil_corrente} → Periodo: €{tot_periodo:,.3f}   |   "
        f"Cumulativo SIL 1→{sil_corrente}: €{tot_cumulativo:,.3f}"
    )

    # Scrivi headers in riga 8
    for c, h in enumerate(P6_IMPORT_HEADERS, 1):
        ws.cell(row=8, column=c, value=h)

    # Cancella dati vecchi (riga 9 in poi, prime 4 colonne)
    if ws.max_row >= 9:
        for row in ws.iter_rows(min_row=9, max_col=4):
            for cell in row:
                cell.value = None

    # Scrivi nuovi dati
    for r, row in enumerate(import_rows, 9):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)


# ─── Funzioni comando ─────────────────────────────────────────────────────────

def _load_all(wb_path: Path) -> Tuple:
    """Carica tutti i dati dall'Excel e restituisce le strutture dati."""
    print(f"  Caricamento: {wb_path.name} ...")
    wb = load_workbook(str(wb_path), read_only=True, data_only=True)

    required = ["MAPPING", "BUDGET da CPM", "INPUT da P6",
                "SIL diretti", "SIL indiretti"]
    for s in required:
        if s not in wb.sheetnames:
            sys.exit(f"[ERRORE] Foglio '{s}' non trovato nel file.")

    print("  Lettura MAPPING ...")
    wbs_to_acts, wbs_to_des, wbs_to_tipo = load_mapping(wb["MAPPING"])

    print("  Lettura BUDGET da CPM ...")
    art_to_cost = load_budget(wb["BUDGET da CPM"])

    print("  Lettura INPUT da P6 (XER) ...")
    act_info = load_p6_tasks(wb["INPUT da P6"])

    print("  Lettura SIL diretti ...")
    sil_dir = load_sil_diretti(wb["SIL diretti"])

    print("  Lettura SIL indiretti ...")
    sil_ind = load_sil_indiretti(wb["SIL indiretti"])

    wb.close()

    sil_records = sil_dir + sil_ind
    sil_max = max((r["sil"] for r in sil_records), default=1)

    print(f"  → {len(wbs_to_acts)} WBS mappate | "
          f"{len(act_info)} attività P6 | "
          f"{len(sil_dir)} SIL diretti | "
          f"{len(sil_ind)} SIL indiretti | "
          f"SIL max = {sil_max}")

    return wbs_to_acts, wbs_to_des, wbs_to_tipo, art_to_cost, act_info, sil_records, sil_max


def cmd_analizza(wb_path: Path, sil_corrente: Optional[int] = None) -> None:
    print("\n[ANALISI]")
    wbs_to_acts, wbs_to_des, wbs_to_tipo, art_to_cost, act_info, sil_records, sil_max = \
        _load_all(wb_path)

    sil_cur = sil_corrente if sil_corrente else sil_max
    print(f"  SIL corrente usato per i report: {sil_cur}")

    print("  Generazione BRIDGE_SIL in memoria ...")
    bridge = generate_bridge(
        sil_records, wbs_to_acts, wbs_to_des, wbs_to_tipo, act_info, art_to_cost
    )
    print(f"  → {len(bridge)} righe BRIDGE_SIL generate")

    show_riepilogo(sil_records, bridge)
    show_alert(bridge, act_info, sil_cur)


def cmd_rigenera(wb_path: Path, sil_corrente: Optional[int] = None) -> None:
    print("\n[RIGENERA]")
    wbs_to_acts, wbs_to_des, wbs_to_tipo, art_to_cost, act_info, sil_records, sil_max = \
        _load_all(wb_path)

    sil_cur = sil_corrente if sil_corrente else sil_max
    print(f"  SIL corrente: {sil_cur}")

    print("  Generazione BRIDGE_SIL ...")
    bridge = generate_bridge(
        sil_records, wbs_to_acts, wbs_to_des, wbs_to_tipo, act_info, art_to_cost
    )
    print(f"  → {len(bridge)} righe")

    print("  Generazione P6_IMPORT_PULITO ...")
    p6_import = generate_p6_import(bridge, sil_cur, act_info)
    print(f"  → {len(p6_import)} attività")

    print(f"  Apertura Excel in scrittura: {wb_path.name} ...")
    wb = load_workbook(str(wb_path))

    for sheet in ["BRIDGE_SIL", "P6_IMPORT_PULITO"]:
        if sheet not in wb.sheetnames:
            sys.exit(f"[ERRORE] Foglio '{sheet}' non trovato.")

    print("  Scrittura BRIDGE_SIL ...")
    _write_bridge(wb["BRIDGE_SIL"], bridge)

    print("  Scrittura P6_IMPORT_PULITO ...")
    _write_p6_import(wb["P6_IMPORT_PULITO"], p6_import, sil_cur, sil_max)

    backup = wb_path.with_stem(wb_path.stem + "_BACKUP")
    import shutil
    shutil.copy2(wb_path, backup)
    print(f"  Backup salvato: {backup.name}")

    wb.save(str(wb_path))
    print(f"  ✓ File aggiornato: {wb_path.name}")

    show_riepilogo(sil_records, bridge)
    show_alert(bridge, act_info, sil_cur)


def cmd_export(
    wb_path: Path,
    output: Optional[Path] = None,
    sil_corrente: Optional[int] = None,
) -> None:
    print("\n[EXPORT CSV]")
    wbs_to_acts, wbs_to_des, wbs_to_tipo, art_to_cost, act_info, sil_records, sil_max = \
        _load_all(wb_path)

    sil_cur = sil_corrente if sil_corrente else sil_max
    print(f"  SIL corrente: {sil_cur}")

    print("  Generazione BRIDGE_SIL ...")
    bridge = generate_bridge(
        sil_records, wbs_to_acts, wbs_to_des, wbs_to_tipo, act_info, art_to_cost
    )

    print("  Generazione P6_IMPORT_PULITO ...")
    p6_import = generate_p6_import(bridge, sil_cur, act_info)

    out_path = output or wb_path.parent / f"P6_IMPORT_SIL{sil_cur:02d}.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(P6_IMPORT_HEADERS)
        writer.writerows(p6_import)

    print(f"  ✓ CSV esportato ({len(p6_import)} righe): {out_path}")

    # Riepilogo a schermo
    tot_per = sum(r[2] for r in p6_import)
    tot_cum = sum(r[3] for r in p6_import)
    print(f"  Periodo (SIL {sil_cur}):   € {tot_per:>15,.2f}")
    print(f"  Cumulativo (SIL 1→{sil_cur}): € {tot_cum:>15,.2f}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="CPM→P6 Bridge — COSEDIL S.p.A. · PMO",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "comando",
        choices=["analizza", "rigenera", "export"],
        help="Operazione da eseguire",
    )
    parser.add_argument(
        "--file", "-f",
        type=Path,
        default=DEFAULT_FILE,
        help=f"Percorso del file Excel Bridge (default: {DEFAULT_FILE.name})",
    )
    parser.add_argument(
        "--sil",
        type=int,
        default=None,
        help="Numero SIL corrente (default: max trovato nei dati)",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help="Percorso output CSV (solo per 'export')",
    )

    args = parser.parse_args()

    if not args.file.exists():
        sys.exit(f"[ERRORE] File non trovato: {args.file}")

    if args.comando == "analizza":
        cmd_analizza(args.file, args.sil)
    elif args.comando == "rigenera":
        cmd_rigenera(args.file, args.sil)
    elif args.comando == "export":
        cmd_export(args.file, args.output, args.sil)


if __name__ == "__main__":
    main()
